# fees/utils.py
from datetime import date
from django.db import transaction
from fees.models import Invoice, FeeStructure
from students.models import Student
import io, zipfile
from django.core import signing
from reportlab.lib.pagesizes import inch
from datetime import datetime

RECEIPT_TOKEN_MAX_AGE = 60 * 60 * 24 * 7  # 7 days (adjust as needed)

def make_receipt_token(payments, school_id):
    """
    Generate a signed token for receipt downloads.
    """
    payment_ids = [p.id for p in payments]
    data = {
        "s": school_id,
        "p": payment_ids,
    }
    return signing.dumps(data)

def parse_receipt_token(token):
    """
    Parse a receipt token like '3-314-315-00d9dc' into components.
    Returns (school_id, payment_ids, random_suffix).
    Raises ValueError if invalid.
    """
    parts = token.split("-")
    if len(parts) < 3:
        raise ValueError("Invalid token format")

    school_id = parts[0]
    random_suffix = parts[-1]
    try:
        payment_ids = [int(p) for p in parts[1:-1]]
    except ValueError:
        raise ValueError("Invalid payment IDs in token")

    return int(school_id), payment_ids, random_suffix

def is_recurring_fee(fee: FeeStructure) -> bool:
    """
    Decide if a FeeStructure is recurring.
    Registration is a one-time fee, all others are recurring by default.
    """
    return fee.name != FeeStructure.REGISTRATION


def generate_invoices_for_school(school):
    """
    Generate invoices for all students in a school.
    - Creates an opening balance invoice (with correct starting_billing_month).
    - Invoices all missed months up to today for recurring fees.
    - Prevents duplicate registration invoices.
    """
    today = date.today()
    new_invoices = []
    count = 0


    # 🔹 Get all students linked to this school
    students = Student.objects.filter(
        division__school=school
    ).prefetch_related("fee_structures")

    for student in students:
        # 🔹 Skip students with no fees or no billing setup
        if not student.fee_structures.exists():
            print(f"⚠️ Skipping {student.full_name} (No fee structures)")
            continue
        if not student.next_payment_date:
            print(f"⚠️ Skipping {student.full_name} (No next payment date)")
            continue

        # 🔹 Create Opening Balance Invoice
        if student.opening_balance and student.opening_balance > 0:
            if not Invoice.objects.filter(
                student=student, status="OPENING_BALANCE"
            ).exists():
                due_date = student.starting_billing_month or today
                billing_month = (student.starting_billing_month or today).replace(day=1)
                new_invoices.append(
                    Invoice(
                        school=school,
                        student=student,
                        fee=None,
                        amount_due=student.opening_balance,
                        due_date=due_date,
                        billing_month=billing_month,
                        status="OPENING_BALANCE",
                        description="Opening Balance"
                    )
                )
                print(f"💰 Created opening balance invoice for {student.full_name}")

        # 🔹 Determine earliest billing date
        next_payment = max(
            filter(None, [student.next_payment_date, student.starting_billing_month]),
            default=None
        )
        if not next_payment or next_payment > today:
            continue

        # 🔹 Generate invoices until caught up
        created_for_student = False
        while next_payment and next_payment <= today:
            billing_month = next_payment.replace(day=1)

            for fee in student.fee_structures.all():
                # Skip duplicate REGISTRATION invoices
                if not is_recurring_fee(fee):
                    if Invoice.objects.filter(student=student, fee=fee).exists():
                        continue

                # Skip if already billed this month
                if Invoice.objects.filter(
                    student=student, fee=fee, billing_month=billing_month
                ).exists():
                    continue

                # Create new invoice
                new_invoices.append(
                    Invoice(
                        school=school,
                        student=student,
                        fee=fee,
                        amount_due=fee.amount,
                        due_date=next_payment,
                        billing_month=billing_month,
                        status="UNPAID",
                    )
                )
                count += 1
                created_for_student = True

            # 🔹 Advance billing cycle
            student.next_payment_date = student.calculate_next_payment_date()
            next_payment = student.next_payment_date

        # 🔹 Save only if invoices were created
        if created_for_student:
            student.payment_status = "PENDING"
            student.save(update_fields=["next_payment_date", "payment_status"])

    # 🔹 Bulk insert invoices
    if new_invoices:
        with transaction.atomic():
            Invoice.objects.bulk_create(new_invoices)

    return count


#For all Exports

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.http import HttpResponse


def export_invoices_to_excel(queryset):
    wb = Workbook()
    ws = wb.active
    ws.append(["Invoice ID", "Student", "Amount Due", "Amount Paid", "Status", "Billing Month"])
    for invoice in queryset:
        ws.append([
            invoice.id,
            invoice.student.full_name,
            invoice.amount_due,
            invoice.amount_paid,
            invoice.status,
            invoice.billing_month
        ])
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment; filename=invoices.xlsx"
    wb.save(response)
    return response

def export_multiple_payment_receipts_pdf(payments):
    """Generate a narrow, receipt-style PDF summarizing multiple payments."""
    # Narrow receipt size (width=3 inches, dynamic height)
    receipt_width = 3 * inch
    line_height = 12
    num_lines = 15 + len(payments)  # Adjust for content
    receipt_height = num_lines * line_height

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=receipt_summary.pdf"

    c = canvas.Canvas(response, pagesize=(receipt_width, receipt_height))
    c.setFont("Courier", 10)

    y = receipt_height - line_height

    def line(text="", indent=0):
        nonlocal y
        c.drawString(10 + indent, y, text)
        y -= line_height

    # === Header ===
    line("SCHOOL PAYMENT RECEIPT")
    line("-------------------------------")
    line(f"DATE: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    line(f"STUDENT: {payments[0].invoice.student.full_name}")
    line("-------------------------------")

    # === Payments ===
    total_paid = 0
    for p in payments:
        total_paid += p.amount
        line(f"INV#{p.invoice.id}  {p.amount:.2f} Br.")

    line("-------------------------------")
    line(f"TOTAL PAID: {total_paid:.2f} Br.")
    line(f"PAYMENTS: {len(payments)}")
    line(f"METHOD(S): {', '.join(set(p.method for p in payments))}")
    line("-------------------------------")
    line("THANK YOU!")

    c.showPage()
    c.save()
    return response

def export_separate_payment_receipts_zip(payments):
    """Creates a ZIP file with separate PDFs for each payment."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zip_file:
        for payment in payments:
            pdf_buffer = io.BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=A4)
            c.drawString(100, 800, f"Receipt for Payment #{payment.id}")
            c.drawString(100, 780, f"Student: {payment.invoice.student.full_name}")
            c.drawString(100, 760, f"Invoice: {payment.invoice.id}")
            c.drawString(100, 740, f"Amount Paid: {payment.amount} Br.")
            c.drawString(100, 720, f"Paid On: {payment.paid_on}")
            c.drawString(100, 700, f"Method: {payment.method}")
            c.showPage()
            c.save()
            zip_file.writestr(f"receipt_{payment.id}.pdf", pdf_buffer.getvalue())

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = "attachment; filename=receipts.zip"
    return response

def export_students_to_excel(qs):
    wb = Workbook()
    ws = wb.active
    ws.append(["Student", "Total Paid", "Total Unpaid", "Invoices Count"])
    for student in qs:
        ws.append([
            student.full_name,
            student.total_paid or 0,
            student.total_unpaid or 0,
            student.invoices_count
        ])
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment; filename=students.xlsx"
    wb.save(response)
    return response

def export_student_invoices_to_excel(student):
    invoices = student.invoices.all()
    wb = Workbook()
    ws = wb.active
    ws.append(["Invoice ID", "Date", "Status", "Amount Due", "Amount Paid", "Billing Month"])
    for invoice in invoices:
        ws.append([
            invoice.id,
            invoice.created_at.strftime("%Y-%m-%d"),
            invoice.status,
            invoice.amount_due,
            invoice.amount_paid or 0,
            invoice.billing_month
        ])
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = f'attachment; filename="{student.full_name}_invoices.xlsx"'
    wb.save(response)
    return response


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def export_student_invoices_to_pdf(student):
    invoices = student.invoices.all()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{student.full_name}_invoices.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Invoices for {student.full_name}", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Invoice ID", "Date", "Status", "Amount Due", "Amount Paid"]]
    for invoice in invoices:
        data.append([
            str(invoice.id),
            invoice.created_at.strftime("%Y-%m-%d"),
            invoice.status,
            f"{invoice.amount_due} Br",
            f"{invoice.amount_paid or 0} Br"
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

from django.utils.timezone import localtime

def export_payments_to_excel(payments):
    """Export selected payments as an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Header
    headers = ["Payment ID", "Invoice ID", "Student", "Fee Type", "Amount Paid", "Paid On", "Method"]
    ws.append(headers)

    # Rows
    for p in payments:
        ws.append([
            p.id,
            p.invoice.id if p.invoice else "",
            p.invoice.student.full_name if p.invoice and p.invoice.student else "",
            p.invoice.fee.name if p.invoice and p.invoice.fee else "Opening Balance",
            float(p.amount),
            localtime(p.paid_on).strftime('%Y-%m-%d %H:%M'),
            p.method,
        ])

    # Return response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
    wb.save(response)
    return response



def generate_payments_pdf(payments):
    """Generate a professional-looking PDF report of payments."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Payments Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Table Data
    data = [["ID", "Student", "Amount", "Method", "Paid On"]]
    for p in payments:
        data.append([
            str(p.id),
            p.invoice.student.full_name,
            f"{p.amount:.2f}",
            p.method,
            localtime(p.paid_on).strftime("%Y-%m-%d"),
        ])

    # Table Styling
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),  # Blue header
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="payments_report.pdf"'
    return response


def generate_payments_excel(payments):
    """Generate an Excel report of payments."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["ID", "Student", "Amount", "Method", "Paid On"]
    ws.append(headers)

    for p in payments:
        ws.append([
            p.id,
            p.invoice.student.full_name,
            p.amount,
            p.method,
            localtime(p.paid_on).strftime("%Y-%m-%d"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="payments_report.xlsx"'
    wb.save(response)
    return response